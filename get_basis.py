from openpyxl import load_workbook
import pickle

#Выбор рабочего листа
file = 'samples/Котировки февраль 26.xlsx'
wb = load_workbook(file, data_only=True)
ws = wb['создано скриптом']

producers_list = ('АМТ', 'ММК', 'Северсталь', 'Аша', 'ЕВРАЗ', 'МЗ Балаково',
                  'НЛМК-Урал', 'Caspian Steel')
basis_list = []

for cell in ws['A']:
    if cell.value in producers_list:
        for c in ws[cell.row]:
            if c.value not in producers_list and c.value != 'типоразмер':
                basis_list.append(c.value)

basis_set = set(basis_list)
basis_set.discard(None)
basis_set.discard('скидка от прайса 24000')
basises_tuple = tuple(sorted(basis_set))

for basis in basises_tuple:
    print(basis)

with open("samples/basises.pkl", "wb") as f:
    pickle.dump(basises_tuple, f)


