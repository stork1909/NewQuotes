from openpyxl import load_workbook

def create_entries():
    #Выбор рабочего листа
    file = 'samples/Котировки февраль 26.xlsx'
    wb = load_workbook(file, data_only=True)
    ws = wb['создано скриптом']

    producers_list = ('АМТ', 'ММК', 'Северсталь', 'Аша', 'ЕВРАЗ', 'МЗ Балаково',
                      'НЛМК-Урал', 'Caspian Steel')

    product_list = ('арматура', 'х/к лист', 'г/к лист', 'Угол', 'Угол ', 'Угол н/п', 'Швеллер', 'Двутавр')

    # Текущий производитель
    current_producer = ''
    # Текущий продукт
    current_product = ''
    # Текущий размер
    current_size = ''
    # Содержит списки базисов и цен
    prices = []
    # Список словарей, которые содержат значения производителя, товара, размера, базиса и цены
    data = []

    # Перебор ячеек первого столбца
    for cell in ws['A']:
        current_row = cell.row
        current_column = cell.column-1
        current_producer = current_producer
        # Проверяем соответствует ли запись в ячейки одному из производителей
        if cell.value in producers_list:
            prices = []
            head = []
            current_producer = cell.value
            current_product = ws[current_row+1][current_column].value
            # Если соответствует перебираем текущий ряд
            for c in ws[current_row]:
                # пустые ячейки в ряду пропускаем, а также пропускаем скидку от прайса Евраза
                if not c.value or 'скидка от прайса' in c.value:
                    continue
                # заполненные ячейки добавляем в список базисов
                else:
                    head.append(c.value)
            # добавляем список базисов (head) в список базисов и цен (prices) и удаляем ненужные вхождения
            # (производителя и слово 'типоразмер')
            prices.append(head)
            del prices[0][0]
            del prices[0][0]
        # Проверяем соответствует ли запись в ячейки одному из продуктов
        elif cell.value in product_list:
            current_product = cell.value
            current_size = ws[current_row][current_column+1].value
            values = []
            # Берем длину заголовка с добавлением удаленных ранее значений и заполняем
            # список цен(values) начиная с 3ей позиции пустыми значениями
            for i in range(2, len(head)+2):
                values.append(None)
            # Если в списке базисов и цен (prices) хранится прошлое значение цен, удаляем его
            if len(prices) > 1:
                del prices[1]
            # И добавляем новое значение цен
            prices.append(values)
        # Если в списке базисов и цен (prices) хранятся и базисы и цены складываем
        # окончательный словарь
        if len(prices) > 1:
            price = dict(zip(prices[0], prices[1]))
            price = {'producer': current_producer, 'product': current_product, 'size': current_size, **price}
            data.append(price)

    #for price in data:
        #print(price)

    return data