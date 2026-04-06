import tkinter as tk
from get_date import get_custom_date
from openpyxl import Workbook

root = tk.Tk()
root.title("Курсы валют")
root.geometry("500x250")

usd_kzt = None
usd_rur = None
rur_kzt = None
date = get_custom_date()
currencies = {0:['курс доллар тенге', usd_kzt, date],
              1:['курс доллар рубль', usd_rur, date],
              2:['курс рубль тенге', rur_kzt, date]
              }

for i, currency in currencies.items():
    tk.Label(root, text=currency[0]).grid(row=i, column=0, padx=10, pady=5, sticky="w")
    currency[1] = tk.Entry(root)
    currency[1].grid(row=i, column=1, padx=10, pady=5, sticky="w")
    date = tk.Entry(root)
    date.grid(row=i, column=2, padx=10, pady=5, sticky="w")
    date.insert(0, currency[2])

def click():
    filename = f'C:\\Users\\Дмитрий\\PycharmProjects\\NewQuotes\\samples\\Тест1.xlsx'
    wb = Workbook()
    ws = wb.active
    for i, currency in currencies.items():
        ws[f'A{i+1}'].value = currency[0]
        ws[f'B{i+1}'].value = currency[1].get()
        ws[f'C{i+1}'].value = currency[2]
        wb.save(filename)
    root.destroy()

button = tk.Button(root, text='Выгрузить', command=click)
button.grid(row=3, column=0, padx=10, pady=5, sticky="w")

root.mainloop()