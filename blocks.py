import pickle
from openpyxl import load_workbook

#Выбор рабочего листа
file = 'samples/Котировки февраль 26.xlsx'
wb = load_workbook(file, data_only=True)
ws = wb['создано скриптом']

producers_list = ('АМТ', 'ММК', 'Северсталь', 'Аша', 'ЕВРАЗ', 'МЗ Балаково',
                  'НЛМК-Урал', 'Caspian Steel')

product_list = ('арматура', 'х/к лист', 'г/к лист', 'Угол', 'Угол ', 'Угол н/п', 'Швеллер', 'Двутавр')


def create_heads():
    blocks = []
    for cell in ws['A']:
        block = []
        if cell.value in producers_list:
            curr_row = cell.row
            for c in ws[curr_row]:
                if c.value == None or 'скидка от прайса' in c.value:
                    pass
                elif c.value in producers_list:
                    pass
                else:
                    block.append(c.value)
            block.insert(0,'производитель')
            block.insert(1,'продукт')
            blocks.append(block)
    return blocks

def create_bodies():
    blocks = []
    entries = []
    producer = None
    for cell in ws['A']:
        entry = []
        curr_row = cell.row
        curr_col = cell.column
        if cell.value in producers_list:
            producer = cell.value
        elif cell.value in product_list:

            for index, c in enumerate(ws[curr_row],1):
                if c.value:
                    if index < 3:
                        entry.append(c.value)
                    else:
                        entry.append(None)
                else:
                    pass
        else:
            pass
        if entry:
            entry.insert(0,producer)
            entries.append(entry)
        if cell.value in producers_list and entries:
            blocks.append(entries)
            entries = []
    blocks.append(entries)
    return blocks

def create_blocs():
    heads = create_heads()
    bodies = create_bodies()

    blocks = []
    for i,body in enumerate(bodies,0):
        block = []
        for entry in body:
            record = dict(zip(heads[i], entry))
            block.append(record)
        blocks.append(block)

    return blocks

def create_forms_data():
    blocks = create_blocs()
    forms = []
    for block in blocks:
        form_data = []
        form_data.append(block[0]['производитель'])
        form_data.append(block[0]['продукт'])
        form_data.append(block[-1]['типоразмер'])
        forms.append(form_data)
    return forms

forms = create_forms_data()
for form in forms:
    print(form)

with open("samples/forms_v2", "wb") as f:
    pickle.dump(forms, f)